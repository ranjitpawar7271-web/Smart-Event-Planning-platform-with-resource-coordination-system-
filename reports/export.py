"""
Turns a report dict (see data.py) into a downloadable file. One function
per format, all reading the exact same structure, so a report only has to
be "shaped" correctly once and every export format follows for free.
"""
import csv
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils.text import slugify


def _filename(report, ext):
    return f"{slugify(report['title'])}.{ext}"


def _is_number(value):
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


# Characters openpyxl forbids in a worksheet title (also disallows a
# leading/trailing apostrophe and a 31-char length cap, both handled below).
_INVALID_SHEET_TITLE_CHARS = r'\/*?:[]'


def _safe_sheet_title(title):
    """Sanitize a report title into a valid openpyxl worksheet name.

    Report titles are free text like "Revenue / Expense / Profit-Loss
    Report" — the '/' there isn't just cosmetic, openpyxl raises
    ValueError on any of \\/*?:[] in a sheet title, so a raw slice of the
    title (as before) crashes the XLSX export for any report whose title
    contains one of those characters.
    """
    cleaned = ''.join(
        ' ' if ch in _INVALID_SHEET_TITLE_CHARS else ch for ch in title
    )
    cleaned = ' '.join(cleaned.split())  # collapse the double spaces left behind
    cleaned = cleaned.strip("'")  # also invalid as first/last character
    return (cleaned[:31] or 'Report')


# --- CSV --------------------------------------------------------------

def csv_response(report):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "csv")}"'

    writer = csv.writer(response)
    writer.writerow([report['title']])
    writer.writerow([report['subtitle']])
    writer.writerow([])

    if report['summary']:
        writer.writerow(['Summary'])
        for label, value in report['summary']:
            writer.writerow([label, value])
        writer.writerow([])

    writer.writerow(report['columns'])
    for row in report['rows']:
        writer.writerow(row)
    return response


# --- Excel (openpyxl) --------------------------------------------------

def xlsx_response(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(report['title'])

    ws.append([report['title']])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([report['subtitle']])
    ws.append([])

    if report['summary']:
        ws.append(['Summary'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for label, value in report['summary']:
            ws.append([label, value])
        ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append(report['columns'])
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx in range(1, len(report['columns']) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in report['rows']:
        ws.append(row)

    for col_idx, column_name in enumerate(report['columns'], start=1):
        max_len = len(str(column_name))
        for row in report['rows']:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        column_letter = ws.cell(row=header_row_idx, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = min(max_len + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "xlsx")}"'
    return response


# --- PDF (reportlab platypus) -------------------------------------------

def pdf_response(report):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), title=report['title'],
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(report['title'], styles['Title']),
        Paragraph(report['subtitle'], styles['Normal']),
        Spacer(1, 8),
    ]

    if report['summary']:
        summary_data = [[str(label), str(value)] for label, value in report['summary']]
        summary_table = Table(summary_data, colWidths=[60 * mm, 60 * mm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (-1, -1), 0.25, colors.HexColor('#333333')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 12))

    # Note: amounts are prefixed "Rs." rather than the Rupee sign. ReportLab's
    # built-in Helvetica (base-14 Type1) font doesn't include the Rupee glyph
    # and would render it as a solid black box — same class of issue as the
    # Unicode-subscript pitfall the pdf skill warns about.
    currency_cols = report.get('currency_columns', set())
    table_data = [report['columns']]
    for row in report['rows']:
        formatted_row = []
        for idx, cell in enumerate(row):
            if idx in currency_cols and _is_number(cell):
                formatted_row.append(f"Rs. {float(cell):,.2f}")
            else:
                formatted_row.append(str(cell))
        table_data.append(formatted_row)

    if len(table_data) > 1:
        data_table = Table(table_data, repeatRows=1)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f4f4')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(data_table)
    else:
        story.append(Paragraph("No data available for this report.", styles['Normal']))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "pdf")}"'
    return response
